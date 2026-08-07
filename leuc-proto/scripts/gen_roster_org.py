#!/usr/bin/env python3
# AI-GEN-BEGIN
"""从来酷科技通讯录 Excel 生成 leuc-proto/data/roster_org.json（无限级部门路径）。

用法：
  python3 scripts/gen_roster_org.py
  python3 scripts/gen_roster_org.py "/path/to/来酷科技通讯录 (3).xlsx"
"""
from __future__ import annotations

import html
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("需要 openpyxl：pip install openpyxl") from e

try:
    from pypinyin import lazy_pinyin
except ImportError:
    lazy_pinyin = None

ROOT = Path(__file__).resolve().parents[1]
# AI-GEN-BEGIN
DEFAULT_XLSX = ROOT.parent / "ziliao" / "组织人员" / "来酷科技通讯录 (3).xlsx"
# AI-GEN-END
OUT = ROOT / "data" / "roster_org.json"
DEMO_USERS = {
    "zhangsan", "lisi", "wangqiang", "zhaomin", "admin", "qianqi",
    "zhangsan1", "zhouba", "wujiu", "zhangcai", "sunli", "liufang",
    "huangwei", "dengjie",
}


def clean_text(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"</?h>", "", s)
    s = re.sub(r"<[^>]+>", "", s)
    s = html.unescape(s)
    return re.sub(r"\s+", " ", s).strip()


def to_py(name: str) -> str:
    name = clean_text(name)
    if not name:
        return "user"
    if lazy_pinyin:
        return "".join(lazy_pinyin(name)).lower()
    return "user"


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"找不到通讯录：{xlsx}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["通讯录明细"]
    rows = list(ws.iter_rows(values_only=True))[1:]

    path_set: OrderedDict[str, None] = OrderedDict()
    people = []
    seen = set()

    def add_path(path: str) -> str:
        parts = [clean_text(p) for p in path.split("/") if clean_text(p)]
        if not parts:
            parts = ["来酷科技"]
        if parts[0] != "来酷科技":
            parts = ["来酷科技"] + [p for p in parts if p != "来酷科技"]
        for i in range(len(parts)):
            path_set["/".join(parts[: i + 1])] = None
        return "/".join(parts)

    for r in rows:
        if not r or not r[1]:
            continue
        name = clean_text(r[1])
        empno = clean_text(r[2]) if r[2] is not None else ""
        title = clean_text(r[3])
        email = clean_text(r[4]).lower() if r[4] else ""
        city = clean_text(r[5])
        dept = clean_text(r[6])
        path = clean_text(r[7]) or (f"来酷科技/{dept}" if dept else "来酷科技")
        path = add_path(path)
        emp_type = clean_text(r[10]) or "正式员工"
        key = empno if empno else f"{name}|{path}"
        if key in seen:
            continue
        seen.add(key)
        people.append(
            {
                "name": name,
                "empno": empno,
                "title": title,
                "email": email or None,
                "city": city or None,
                "dept_path": path,
                "person_type": "external" if emp_type == "外包" else "internal",
                "emp_type": emp_type,
            }
        )

    for r in list(wb["部门结构"].iter_rows(values_only=True))[1:]:
        if r and r[5]:
            add_path(clean_text(r[5]))

    paths = sorted(path_set.keys(), key=lambda p: (p.count("/"), p))
    depts = []
    path_to_id = {}
    for i, p in enumerate(paths, 1):
        parent = "/".join(p.split("/")[:-1]) if "/" in p else None
        depts.append({"id": i, "name": p.split("/")[-1], "path": p, "parent_path": parent})
        path_to_id[p] = i

    # AI-GEN-BEGIN
    # 登录账号 = 姓名拼音全拼；冲突加 2、3…（工号仍写入 itcode）
    used = set(DEMO_USERS)
    for p in people:
        base = to_py(p["name"]) or "user"
        base = re.sub(r"[^a-z0-9_]", "", base.lower()) or "user"
        if base[0].isdigit():
            base = "u" + base
        uname, n = base, 2
        while uname in used:
            uname = f"{base}{n}"
            n += 1
        used.add(uname)
        p["username"] = uname
        p["dept_id"] = path_to_id[p["dept_path"]]
    # AI-GEN-END

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(
            {
                "source": xlsx.name,
                "root": "来酷科技",
                "departments": depts,
                "people": people,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    print(f"wrote {OUT} depts={len(depts)} people={len(people)}")


if __name__ == "__main__":
    main()
# AI-GEN-END
